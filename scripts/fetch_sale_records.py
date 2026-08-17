#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
京华云采商品销售记录爬取脚本
获取指定品类/品牌/型号机型的历史销售数据，按型号、采购单位、采购数量、采购单价、供应商、成交时间建表。

用法：
  # 指定型号的销售记录
  python fetch_sale_records.py --product 开天M70d --out 开天M70d销售记录.xlsx

  # 指定品牌的销售记录
  python fetch_sale_records.py --brand 联想 --out 联想销售记录.xlsx

  # 指定品类的销售记录
  python fetch_sale_records.py --category 台 --out 台式机销售记录.xlsx

筛选说明：
  --category：品类键（台/笔/服/印），逗号分隔
  --brand：品牌名，模糊匹配
  --product：产品关键词，按 skuName 包含匹配

API 说明：
  - 商品搜索（获取 skuId）：mkt-bjzc.zhongcy.com 的 querySkuListFromEs
  - 销售记录（获取成交）：shop-bjzc.zhongcy.com 的 querySkuSaleRecord（GET + query 参数）
"""
import argparse
import time
import requests
from collections import defaultdict

# 两个域名：mkt 查商品，shop 查销售
MKT_BASE = "https://mkt-bjzc.zhongcy.com"
SHOP_BASE = "https://shop-bjzc.zhongcy.com"
SEARCH_URL = MKT_BASE + "/proxy/trade-service/mall/search/querySkuListFromEs"
AGENT_URL = MKT_BASE + "/proxy/trade-service/mall/search/querySkuAgentListFromEs"
SALE_URL = SHOP_BASE + "/proxy/trade-service/mall/order/querySkuSaleRecord"
PLATFORM_ID = 20
PUBLISH_TYPE = 10024

CATEGORIES = {
    "台": (1000011, "台式计算机"),
    "笔": (1000014, "便携式计算机"),
    "服": (1000128, "服务器"),
    "印": (1000023, "打印机"),
}
PRINTER_CIDS = [1000023, 1000025, 1000027, 1000029, 1000041]

SESSION = requests.Session()
# 关键：绕过系统代理直连。zhongcy.com 是国内政务平台，直连可达；
# 若走系统代理（如 127.0.0.1:7897）会导致 TLS 握手失败（SSLError/UNEXPECTED_EOF）。
SESSION.trust_env = False
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Content-Type": "application/json",
    "Origin": MKT_BASE,
})


def post_json(url, payload, retries=3):
    last_err = "未知错误"
    for i in range(retries):
        try:
            r = SESSION.post(url, json=payload, timeout=30)
            d = r.json()
            if d.get("code") == "0":
                return d
            if "99999" in str(d.get("code", "")):
                time.sleep(0.5 * (i + 2))
                continue
            return d
        except requests.exceptions.Timeout:
            last_err = "网络超时"
            time.sleep(0.5 * (i + 2))
        except requests.exceptions.ConnectionError:
            last_err = "网络连接失败"
            time.sleep(0.5 * (i + 2))
        except requests.exceptions.SSLError:
            last_err = "SSL/代理问题"
            time.sleep(0.5 * (i + 2))
        except Exception as e:
            last_err = type(e).__name__
            time.sleep(0.5 * (i + 2))
    print(f"⚠️ 请求失败[{last_err}]（已重试 {retries} 次）：{url}，请检查网络或代理")
    return None


def get_json(url, params, retries=3):
    """GET 请求（用于销售记录接口）"""
    last_err = "未知错误"
    for i in range(retries):
        try:
            r = SESSION.get(url, params=params, timeout=30)
            d = r.json()
            return d
        except requests.exceptions.Timeout:
            last_err = "网络超时"
            time.sleep(0.5 * (i + 2))
        except requests.exceptions.ConnectionError:
            last_err = "网络连接失败"
            time.sleep(0.5 * (i + 2))
        except requests.exceptions.SSLError:
            last_err = "SSL/代理问题"
            time.sleep(0.5 * (i + 2))
        except Exception as e:
            last_err = type(e).__name__
            time.sleep(0.5 * (i + 2))
    print(f"⚠️ 请求失败[{last_err}]（已重试 {retries} 次）：{url}，请检查网络或代理")
    return None


def get_cid_list(category_keys):
    if not category_keys:
        cids = []
        for k, (cid, _) in CATEGORIES.items():
            if k == "印":
                cids.extend(PRINTER_CIDS)
            else:
                cids.append(cid)
    else:
        cids = []
        for k in category_keys:
            if k == "印":
                cids.extend(PRINTER_CIDS)
            elif k in CATEGORIES:
                cids.append(CATEGORIES[k][0])
    seen, result = set(), []
    for c in cids:
        if c not in seen:
            seen.add(c)
            result.append(c)
    return result


def cid_to_name(cid):
    for k, (c, name) in CATEGORIES.items():
        if c == cid:
            return name
    return str(cid)


def get_products(cid):
    products = []
    page, page_size = 1, 100
    while True:
        payload = {
            "queryPage": {"platformId": PLATFORM_ID, "pageSize": page_size, "pageNum": page},
            "homeType": 110,
            "isAggregation": False,
            "isHometypes": True,
            "cid": cid,
            "businessType": 1,
            "publishType": PUBLISH_TYPE,
        }
        d = post_json(SEARCH_URL, payload)
        if not d or not d.get("data"):
            break
        rl = d["data"]["itemList"]["resultList"]
        if not rl:
            break
        for it in rl:
            products.append({
                "skuId": it["skuId"],
                "skuName": it.get("skuName", ""),
                "brandName": (it.get("brandNameCh") or "") + (it.get("brandNameEn") or ""),
                "category": cid_to_name(cid),
            })
        total = int(d["data"].get("totalCount", 0) or 0)
        if page * page_size >= total:
            break
        page += 1
        time.sleep(0.2)
    return products


def get_sale_records(sku_id):
    """获取某商品的销售记录（分页），返回记录列表"""
    records = []
    page = 1
    page_size = 100  # 上限 100（推测，与平台一致）
    while True:
        params = {
            "pageNum": page,
            "pageSize": page_size,
            "platformId": PLATFORM_ID,
            "skuId": sku_id,
        }
        d = get_json(SALE_URL, params)
        if not d or d.get("code") != "0":
            break
        data = d.get("data") or {}
        # 实测：销售记录在 data.result，总数在 data.totalCount，总页数 data.totalPageCount
        rl = data.get("result") or data.get("resultList") or []
        if not rl:
            break
        records.extend(rl)
        total = int(data.get("totalCount", 0) or 0)
        total_pages = int(data.get("totalPageCount", 0) or 0)
        # 用总页数判断是否还有下一页，更可靠
        if total_pages and page >= total_pages:
            break
        if page * page_size >= total:
            break
        page += 1
        time.sleep(0.25)
    return records


def extract_record(item, sku_name, brand, category):
    """从一条销售记录里提取字段（字段名已按实测返回对齐）

    实测返回示例：
    {"organizeId":19231, "organizeName":"北京市平谷区投资促进服务中心本级",
     "shopId":610, "shopName":"北京绿都畅达人才科技发展有限公司",
     "skuNum":1, "sellPrice":5900.00000, "orderTime":"2025-01-21 11:14:49"}
    """
    return {
        "category": category,
        "brand": brand,
        # 销售记录里没有型号字段，用商品的 skuName 补
        "skuName": item.get("skuName") or sku_name,
        "buyerName": item.get("organizeName", ""),
        "quantity": item.get("skuNum", ""),
        "unitPrice": item.get("sellPrice", ""),
        "supplierName": item.get("shopName", ""),
        "saleTime": item.get("orderTime", ""),
    }


def export_excel(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "销售记录"

    headers = ["品类", "品牌", "型号", "采购单位", "采购数量", "采购单价(元)", "供应商", "成交时间"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2F5496")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, 2):
        vals = [
            row["category"], row["brand"], row["skuName"], row["buyerName"],
            row["quantity"], row["unitPrice"], row["supplierName"], row["saleTime"],
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=r, column=col, value=v)

    widths = [14, 12, 30, 28, 10, 12, 26, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"输出：{out_path}（共 {len(rows)} 条销售记录）")


def main():
    parser = argparse.ArgumentParser(description="京华云采商品销售记录爬取")
    parser.add_argument("--category", default="", help="品类键（台/笔/服/印），逗号分隔")
    parser.add_argument("--brand", default="", help="品牌名（如 联想），逗号分隔")
    parser.add_argument("--product", default="", help="产品关键词（如 开天M70d）")
    parser.add_argument("--out", default="销售记录.xlsx", help="输出 Excel 路径")
    args = parser.parse_args()

    cat_keys = [k.strip() for k in args.category.split(",") if k.strip()]
    brands = [b.strip() for b in args.brand.split(",") if b.strip()]
    product_kw = args.product.strip()

    print(f"筛选条件：品类={cat_keys or '全部'}，品牌={brands or '全部'}，产品={product_kw or '全部'}")

    cids = get_cid_list(cat_keys)
    all_products = []
    for cid in cids:
        prods = get_products(cid)
        all_products.extend(prods)
        print(f"cid={cid}（{cid_to_name(cid)}）：{len(prods)} 个商品")

    # 筛选商品
    filtered = []
    for p in all_products:
        if brands and not any(b in p["brandName"] for b in brands):
            continue
        if product_kw and product_kw not in p["skuName"]:
            continue
        filtered.append(p)
    print(f"筛选后 {len(filtered)} 个商品，开始爬取销售记录...")

    all_records = []
    for idx, p in enumerate(filtered, 1):
        records = get_sale_records(p["skuId"])
        for rec in records:
            all_records.append(extract_record(rec, p["skuName"], p["brandName"], p["category"]))
        if idx % 10 == 0:
            print(f"  [{idx}/{len(filtered)}] 已累计 {len(all_records)} 条销售记录")
        time.sleep(0.2)

    print(f"\n爬取完成：{len(all_records)} 条销售记录")
    export_excel(all_records, args.out)


if __name__ == "__main__":
    main()
