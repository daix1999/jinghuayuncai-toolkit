#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
京华云采名单电话填充脚本（通用版）
用法：
  python fill_phones.py <输入xlsx路径> <映射json路径> [--name-col A] [--phone-col D]
                        [--remark-col F] [--out <输出路径>]

列参数支持两种写法：
  - 列号：A / B / D，或 1 / 2 / 4
  - 列名：供应商名称 / 联系电话 / 代理商名称 等（自动在表头行查找，模糊包含匹配）

功能：
  - 读取供应商名单 Excel，找出电话列为空的行
  - 根据映射 JSON（公司名 -> {phone, source, contact}）填充电话
  - 新增"电话来源"列和"可能联系人及身份"列
  - 颜色标注：绿色=API匹配、黄色=公开搜索、橙色=未找到
  - 跑完打印完整摘要（总行数/空号数/已填/未找到/未匹配/跳过），不静默失败
  - 默认输出到"原名_已填充.xlsx"副本，不覆盖原文件
"""
import argparse
import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


def col_to_idx(col):
    """把 'A'/'B' 或数字转换成 openpyxl 列索引（1-based）"""
    if isinstance(col, int):
        return col
    return column_index_from_string(col.strip().upper())


def resolve_col(ws, header_row, col_spec, field_desc):
    """解析列参数：支持列号（A/B/1/2）或列名（供应商名称）。找不到时明确报错并列出真实表头。"""
    if col_spec is None:
        return None
    s = str(col_spec).strip()
    # 纯字母 -> 列号
    if s.isalpha():
        return column_index_from_string(s.upper())
    # 纯数字 -> 列号
    if s.isdigit():
        return int(s)
    # 否则当作列名，在表头行模糊匹配
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and s in str(v):
            return c
    # 找不到，报错并列出真实表头
    headers = [
        str(ws.cell(row=header_row, column=c).value)
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value
    ]
    raise SystemExit(
        f"❌ 找不到「{s}」对应的列（{field_desc}）。\n"
        f"   表头行实际为：{headers}\n"
        f"   请用 --name-col / --phone-col 指定正确的列名或列号后重试。"
    )


def is_empty(val):
    if val is None:
        return True
    s = str(val).strip()
    return s in ('', 'nan', 'None', '无', 'null')


def fill_phones(input_path, mapping, name_col='A', phone_col='D',
                remark_col=None, source_col_name='电话来源',
                contact_col_name='可能联系人及身份', out_path=None):
    wb = load_workbook(input_path)
    ws = wb.active

    # 先按默认 name_col 找表头行（若 name_col 是列名，先用 A 列找表头）
    probe_idx = col_to_idx(name_col) if str(name_col).strip() in ('ABCDEFGHIJKLMNOPQRSTUVWXYZ' or '') else 1
    try:
        probe_idx = col_to_idx(name_col)
    except Exception:
        probe_idx = 1
    header_row = 1
    for r in range(1, 5):
        if ws.cell(row=r, column=probe_idx).value:
            header_row = r
            break

    # 解析各列位置（支持列名），找不到会明确报错
    name_idx = resolve_col(ws, header_row, name_col, '供应商名称列')
    phone_idx = resolve_col(ws, header_row, phone_col, '电话列')
    remark_idx = resolve_col(ws, header_row, remark_col, '备注列') if remark_col else None

    green_fill = PatternFill('solid', fgColor='C6EFCE')   # API 匹配
    yellow_fill = PatternFill('solid', fgColor='FFF2CC')  # 公开搜索
    orange_fill = PatternFill('solid', fgColor='FCE4D6')  # 未找到

    # 确定新增列位置
    max_col = ws.max_column
    source_idx = max_col + 1
    contact_idx = max_col + 2
    ws.cell(row=header_row, column=source_idx, value=source_col_name).font = Font(bold=True)
    ws.cell(row=header_row, column=contact_idx, value=contact_col_name).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(source_idx)].width = 28
    ws.column_dimensions[get_column_letter(contact_idx)].width = 32

    filled = 0
    not_found = 0
    skipped = 0
    unmatched = 0          # 名单里电话为空、但 mapping 里没有的公司
    total_rows = 0         # 名单里的有效数据行数
    empty_phone_rows = 0   # 名单里电话为空的行数

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=name_idx)
        name = str(name_cell.value).strip() if name_cell.value else ''
        if not name or name in ('nan', 'None'):
            continue
        total_rows += 1

        phone_cell = ws.cell(row=row_idx, column=phone_idx)
        if not is_empty(phone_cell.value):
            skipped += 1
            continue
        empty_phone_rows += 1

        # 用规范化后的名称匹配
        key = name.strip()
        if key not in mapping:
            unmatched += 1
            # 不静默：标记为"未匹配"（不填电话，但计入统计并给出提示来源）
            ws.cell(row=row_idx, column=source_idx).value = '未匹配(映射缺失)'
            continue

        entry = mapping[key]
        phone = entry.get('phone', '')
        source = entry.get('source', '')
        contact = entry.get('contact', '')

        if phone:
            phone_cell.value = phone
            phone_cell.fill = yellow_fill if ('搜索' in source or '公开' in source or 'web' in source.lower()) else green_fill
            filled += 1
        else:
            phone_cell.fill = orange_fill
            not_found += 1

        ws.cell(row=row_idx, column=source_idx).value = source or '需进一步获取'
        if contact:
            ws.cell(row=row_idx, column=contact_idx).value = contact

        if remark_idx and contact:
            rc = ws.cell(row=row_idx, column=remark_idx)
            existing = str(rc.value).strip() if rc.value and str(rc.value).strip() not in ('nan', 'None') else ''
            if existing and contact not in existing:
                rc.value = existing + '; ' + contact
            elif not existing:
                rc.value = contact

    if out_path is None:
        out_path = input_path.rsplit('.', 1)[0] + '_已填充.xlsx'
    wb.save(out_path)

    # 完整摘要，让用户一眼看清结果完整性
    print('=' * 40)
    print('填充结果摘要：')
    print(f'  名单有效数据行：{total_rows} 行')
    print(f'  其中电话为空：{empty_phone_rows} 行')
    print(f'  ✅ 已填充电话：{filled} 家')
    print(f'  🟠 未找到电话：{not_found} 家')
    print(f'  ⚠️ 未匹配(映射缺失)：{unmatched} 家')
    print(f'  ➖ 已有电话跳过：{skipped} 家')
    if unmatched > 0:
        print(f'  💡 提示：{unmatched} 家电话为空但映射里没有，请检查映射 JSON 是否覆盖全部空号公司')
    print(f'  输出：{out_path}')
    print('=' * 40)

    return {
        'filled': filled, 'not_found': not_found, 'skipped': skipped,
        'unmatched': unmatched, 'total_rows': total_rows,
        'empty_phone_rows': empty_phone_rows, 'output': out_path,
    }


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='京华云采名单电话填充')
    parser.add_argument('input', help='输入 xlsx 路径')
    parser.add_argument('mapping', help='映射 JSON 路径：{公司名: {phone, source, contact}}')
    parser.add_argument('--name-col', default='A', help='供应商名称列，可用列号(A)或列名(供应商名称)，默认 A')
    parser.add_argument('--phone-col', default='D', help='电话列，可用列号(D)或列名(联系电话)，默认 D')
    parser.add_argument('--remark-col', default=None, help='备注列（可选），用于追加联系人')
    parser.add_argument('--out', default=None, help='输出路径（默认 原名_已填充.xlsx）')
    args = parser.parse_args()

    with open(args.mapping, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    fill_phones(args.input, mapping, args.name_col, args.phone_col,
                args.remark_col, out_path=args.out)
