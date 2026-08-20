#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
京华云采销售数据分析（v4：数据 + 关键订单明细 + 说明型报告）

用法：
    python analyze_sales.py <销售记录.xlsx> [--out 分析报告.xlsx]

输入：fetch_sale_records.py 输出的销售记录 Excel
输出：多 Sheet 分析报告，每个分析 Sheet = 「汇总表 + 📌关键订单明细 + 📌分析说明」，
      订单明细具体到采购单位/供应商/型号/配置/数量/单价/时间，一眼可读。

Sheet 结构：
    销售明细 / 整体分析报告 / 品牌对比分析 / 型号分析 / 配置价格稳定性 /
    单位类型分布 / 多次往来 / 供应商占比 / 月度销售趋势 / TOP采购单位 / 价格分析 / 采购频次

说明：型号拆分为「型号+配置」，价格按配置分组；同配置明显波动（>100元或>2%）单独标注为议价信号。
"""
import argparse
import os
import re
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


def find_col(df, *candidates):
    """在 DataFrame 中按候选列名模糊匹配，返回真实列名"""
    for cand in candidates:
        for c in df.columns:
            if cand in str(c):
                return c
    return None


def classify_org(name):
    """按单位名称关键词分类买方类型（学校 > 医院 > 政府 > 企业 > 其他）"""
    school_kw = ['大学', '学院', '中学', '小学', '学校', '幼儿园', '党校',
                 '职业', '技师', '培训学校', '实验学校']
    hospital_kw = ['医院', '卫生院', '卫生服务中心', '疾控', '口腔',
                   '医疗', '妇幼保健', '康复中心', '卫生所']
    gov_kw = ['人民政府', '政府', '街道', '镇', '乡', '局', '委员会', '委',
              '法院', '检察院', '公安', '税务', '财政', '审计', '纪检',
              '人大', '政协', '机关事务', '管委会', '办公室', '监督管理']
    company_kw = ['有限公司', '集团', '股份', '公司', '厂', '合作社', '商贸', '科技']

    if any(k in name for k in school_kw):
        return '学校/教育'
    if any(k in name for k in hospital_kw):
        return '医院/医疗'
    if any(k in name for k in gov_kw):
        return '政府/机关'
    if any(k in name for k in company_kw):
        return '企业'
    return '其他/事业单位'


def split_model_config(model_str):
    """把型号字符串拆分为 (型号, 配置)。括号内为配置，括号外为型号主体。"""
    s = str(model_str).strip()
    m = re.search(r'[（(](.*?)[)）]', s)
    if m:
        config = m.group(1).strip()
        model = s[:m.start()].strip()
    else:
        config = ''
        model = s
    model = re.sub(r'(台式计算机|便携式计算机|笔记本电脑|笔记本|一体机|显示器|台式机)$', '', model).strip()
    return model, config


def append_insights(ws, start_row, title, insights, col=1):
    """在 sheet 的 start_row 下方写入分析说明文本块"""
    ws.cell(row=start_row, column=col, value=title).font = Font(bold=True, size=12)
    for i, line in enumerate(insights, 1):
        ws.cell(row=start_row + i, column=col, value=line)
    return start_row + len(insights) + 2


def write_analysis_sheet(writer, sheet_name, summary_df, detail_df, insights):
    """写一个分析 sheet：汇总表 + 关键订单明细表 + 分析说明"""
    summary_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    ws = writer.sheets[sheet_name]
    r = len(summary_df) + 3
    if detail_df is not None and len(detail_df) > 0:
        ws.cell(row=r, column=1, value='📌 关键订单明细（采购单位 / 供应商 / 型号 / 配置 / 数量 / 单价 / 时间）').font = Font(bold=True, size=12)
        r += 1
        for c, col in enumerate(detail_df.columns, 1):
            cell = ws.cell(row=r, column=c, value=str(col))
            cell.font = Font(bold=True)
        for i, (_, row) in enumerate(detail_df.iterrows(), 1):
            for c, col in enumerate(detail_df.columns, 1):
                ws.cell(row=r + i, column=c, value=row[col])
        r += len(detail_df) + 2
    else:
        r += 1
    append_insights(ws, r, '📌 分析说明', insights)


def gen_overall_report(df, n, total_qty, total_amt, n_org, n_shop,
                       type_stat, shop_stat, monthly, org_stat, prices,
                       brand_stat, model_stat, freq_stat, model_col, price_stab=None,
                       org_col=None, shop_col=None):
    """生成 Sheet2 整体分析报告（文字为主）"""
    L = []
    L.append('【京华云采销售数据分析报告】')
    L.append('')
    L.append('一、数据概览')
    L.append(f'  本次共分析 {n} 条成交记录，覆盖 {n_org} 家采购单位、{n_shop} 家供应商。')
    L.append(f'  总销量 {total_qty} 台，成交总金额 ¥{total_amt:,.0f}，均价 ¥{prices.mean():,.0f}。')
    L.append('')
    L.append('二、品牌格局')
    if brand_stat is not None and len(brand_stat) > 0:
        top_b = brand_stat.iloc[0]
        L.append(f'  共 {len(brand_stat)} 个品牌，主力品牌为【{top_b.name}】：{int(top_b["销量"])} 台、'
                 f'¥{top_b["金额"]:,.0f}（金额占比 {top_b["金额占比"]:.1f}%）。')
        if len(brand_stat) > 1:
            cr3 = brand_stat['金额占比'].head(3).sum()
            L.append(f'  TOP3 品牌金额集中度 {cr3:.1f}%' +
                     ('，品牌格局集中。' if cr3 > 70 else '，品牌格局较分散。'))
    L.append('')
    L.append('三、买方结构（单位类型）')
    t1 = type_stat.iloc[0]
    L.append(f'  成交最多的类型是【{t1.name}】：{int(t1["销量"])} 台、¥{t1["金额"]:,.0f}（金额占比 {t1["金额占比"]:.1f}%）。')
    for t in ['学校/教育', '医院/医疗']:
        if t in type_stat.index and type_stat.loc[t, '金额占比'] < 5:
            L.append(f'  【{t}】渗透率 <5%，是潜在空白市场。')
    L.append('')
    L.append('四、供应商格局')
    s1 = shop_stat.iloc[0]
    L.append(f'  共 {len(shop_stat)} 家供应商供货，头部为【{s1.name}】：金额占比 {s1["金额占比"]:.1f}%。')
    cr3 = shop_stat['金额占比'].head(3).sum()
    L.append(f'  TOP3 供应商金额集中度 {cr3:.1f}%' +
             ('，较集中（头部有壁垒）。' if cr3 > 60 else '，较分散（有切入空间）。'))
    L.append('')
    L.append('五、价格信号')
    L.append(f'  成交均价 ¥{prices.mean():,.0f}，总体区间 ¥{prices.min():,.0f} - ¥{prices.max():,.0f}。')
    L.append('  ※ 价格差异主要来自配置不同（CPU/内存/硬盘/显卡/显示器），须按「型号+配置」对比价格。')
    if price_stab is not None and len(price_stab) > 0:
        stable_n = int(price_stab['价格状态'].isin(['稳定', '基本稳定(微幅)']).sum())
        L.append(f'  按「型号+配置」对比：{stable_n}/{len(price_stab)} 个组合价格稳定。')
        unstable = price_stab[price_stab['价格状态'] == '⚠️ 有波动']
        if len(unstable) > 0:
            L.append(f'  ⚠️ 同配置价格明显波动的组合 {len(unstable)} 个（议价信号，详见「配置价格稳定性」Sheet）：')
            for _, r in unstable.head(3).iterrows():
                cfg = f'（{r["配置"]}）' if r.get('配置') else ''
                L.append(f'    · {str(r["型号"])[:16]}{cfg}：¥{r["最低价"]:,.0f}-¥{r["最高价"]:,.0f}（波动 {r["波动%"]:.0f}%）')
        else:
            L.append('  未发现同配置价格明显波动，市场定价规范。')
    L.append('')
    L.append('六、趋势与节奏')
    if len(monthly) > 0:
        peak = monthly['销量'].idxmax()
        peak_row = monthly.loc[peak]
        L.append(f'  覆盖 {len(monthly)} 个月，销量峰值在【{peak}】：{int(peak_row["销量"])} 台、¥{peak_row["金额"]:,.0f}。')
        growth = monthly['销量'].pct_change().max()
        if pd.notna(growth) and growth > 0.5:
            L.append(f'  单月最大环比增长 {growth*100:.0f}%，存在明显的采购波峰。')
    L.append('')
    L.append('七、结论与建议')
    for t in ['学校/教育', '医院/医疗']:
        if t in type_stat.index and type_stat.loc[t, '金额占比'] < 5:
            L.append(f'  ▶ 【{t}】类单位渗透低，可作为新客户突破口。')
    if cr3 <= 60:
        L.append('  ▶ 供应商格局分散，暂无垄断者，适合切入供货。')
    if model_col and model_stat is not None and len(model_stat) > 0:
        top_m = model_stat.iloc[0]
        cfg = f'（{top_m["配置"]}）' if top_m.get('配置') else ''
        L.append(f'  ▶ 走量款【{str(top_m["型号"])[:24]}{cfg}】（{int(top_m["销量"])} 台），备货与报价优先覆盖。')
    L.append('  ▶ 具体订单见各分析 Sheet 的「关键订单明细」，可直接用于客户沟通。')
    return L


def analyze(input_path, out_path, quiet=False):
    df = pd.read_excel(input_path)
    if df.empty:
        print('❌ 数据为空，无法分析')
        sys.exit(1)

    org_col = find_col(df, '采购单位')
    shop_col = find_col(df, '供应商')
    num_col = find_col(df, '采购数量', '数量')
    price_col = find_col(df, '采购单价', '单价')
    time_col = find_col(df, '成交时间', '下单时间', '时间')
    model_col = find_col(df, '型号', '商品')
    brand_col = find_col(df, '品牌')
    if not (org_col and shop_col and num_col and price_col and time_col):
        print(f'❌ 缺少必要列。当前列：{list(df.columns)}')
        print('   需要：采购单位 / 供应商 / 采购数量 / 采购单价 / 成交时间')
        sys.exit(1)

    df['采购数量'] = pd.to_numeric(df[num_col], errors='coerce').fillna(0)
    df['采购单价'] = pd.to_numeric(df[price_col], errors='coerce').fillna(0)
    df['成交金额'] = df['采购数量'] * df['采购单价']
    df['成交时间'] = pd.to_datetime(df[time_col], errors='coerce')
    df['月份'] = df['成交时间'].dt.strftime('%Y-%m')
    df['单位类型'] = df[org_col].astype(str).apply(classify_org)
    df['时间串'] = df['成交时间'].dt.strftime('%Y-%m-%d %H:%M')
    if model_col:
        _splits = df[model_col].astype(str).apply(split_model_config)
        df['型号名'] = _splits.apply(lambda x: x[0])
        df['配置'] = _splits.apply(lambda x: x[1])

    n = len(df)
    total_qty = int(df['采购数量'].sum())
    total_amt = df['成交金额'].sum()
    n_org = df[org_col].nunique()
    n_shop = df[shop_col].nunique()

    print('=' * 56)
    print('京华云采销售数据分析（v4：带订单明细）')
    print(f'  记录数：{n} 条 | 采购单位：{n_org} 家 | 供应商：{n_shop} 家')
    print(f'  总销量：{total_qty} 台 | 总金额：¥{total_amt:,.0f}')
    print('=' * 56)

    # 明细提取 helper（统一列）
    def make_detail(mask, limit=None, sort_col='时间串'):
        sub = df[mask]
        if limit:
            sub = sub.sort_values(sort_col).head(limit)
        out = pd.DataFrame({
            '采购单位': sub[org_col], '供应商': sub[shop_col],
            '型号': sub['型号名'] if model_col else '',
            '配置': sub['配置'] if model_col else '',
            '采购数量': sub['采购数量'], '采购单价(元)': sub['采购单价'],
            '成交时间': sub['时间串'],
        })
        return out.reset_index(drop=True)

    # ============ 各维度聚合 ============
    brand_stat = None
    if brand_col:
        brand_stat = df.groupby(brand_col).agg(
            销量=('采购数量', 'sum'), 金额=('成交金额', 'sum'),
            客户数=('采购单位', 'nunique'), 记录数=('采购单位', 'count'),
        ).sort_values('金额', ascending=False)
        brand_stat['金额占比'] = (brand_stat['金额'] / total_amt * 100).round(1)

    model_stat = None
    if model_col:
        gb_cols = ([brand_col] if brand_col else []) + ['型号名', '配置']
        agg_map = {
            '销量': ('采购数量', 'sum'), '金额': ('成交金额', 'sum'),
            '记录数': ('采购单位', 'count'), '均价': ('采购单价', 'mean'),
            '最低价': ('采购单价', 'min'), '最高价': ('采购单价', 'max'),
        }
        model_stat = df.groupby(gb_cols).agg(**agg_map).sort_values('销量', ascending=False).reset_index()
        cols = (['品牌'] if brand_col else []) + ['型号', '配置', '销量', '金额', '记录数', '均价', '最低价', '最高价']
        model_stat.columns = cols
        model_stat['均价'] = model_stat['均价'].round(0)

    price_stab = None
    if model_col:
        gb_cols = ([brand_col] if brand_col else []) + ['型号名', '配置']
        grp = df.groupby(gb_cols)['采购单价']
        price_stab = pd.DataFrame({
            '记录数': grp.count(), '均价': grp.mean().round(0),
            '最低价': grp.min(), '最高价': grp.max(),
            '不同价格数': grp.nunique(),
        }).reset_index()
        price_stab['波动幅度'] = price_stab['最高价'] - price_stab['最低价']
        price_stab['波动%'] = (price_stab['波动幅度'] / price_stab['均价'].replace(0, pd.NA) * 100).round(1)

        def _judge_stab(r):
            if r['不同价格数'] <= 1:
                return '稳定'
            if r['波动幅度'] <= 100 or r['波动%'] <= 2:
                return '基本稳定(微幅)'
            return '⚠️ 有波动'

        price_stab['价格状态'] = price_stab.apply(_judge_stab, axis=1)
        price_stab = price_stab.sort_values('波动幅度', ascending=False).reset_index(drop=True)
        cols = (['品牌'] if brand_col else []) + ['型号', '配置', '记录数', '均价', '最低价', '最高价',
                                                  '不同价格数', '波动幅度', '波动%', '价格状态']
        price_stab.columns = cols

    type_stat = df.groupby('单位类型').agg(
        采购单位数=('采购单位', 'nunique'), 销量=('采购数量', 'sum'),
        金额=('成交金额', 'sum'), 记录数=('采购单位', 'count'),
    ).sort_values('金额', ascending=False)
    type_stat['金额占比'] = (type_stat['金额'] / total_amt * 100).round(1)

    pair = df.groupby([org_col, shop_col]).size().reset_index(name='次数')
    pair['销量'] = df.groupby([org_col, shop_col])['采购数量'].sum().values
    pair['金额'] = df.groupby([org_col, shop_col])['成交金额'].sum().values
    pair.columns = ['采购单位', '供应商', '次数', '销量', '金额']
    repeat = pair[pair['次数'] > 1].sort_values('次数', ascending=False)

    shop_stat = df.groupby(shop_col).agg(
        客户数=('采购单位', 'nunique'), 销量=('采购数量', 'sum'),
        金额=('成交金额', 'sum'),
    ).sort_values('金额', ascending=False)
    shop_stat['金额占比'] = (shop_stat['金额'] / total_amt * 100).round(1)
    shop_stat['销量占比'] = (shop_stat['销量'] / total_qty * 100).round(1)

    monthly = df.groupby('月份').agg(
        销量=('采购数量', 'sum'), 金额=('成交金额', 'sum'),
        记录数=('采购单位', 'count'),
    ).sort_index()
    monthly['销量环比%'] = monthly['销量'].pct_change().mul(100).round(1)

    org_stat = df.groupby(org_col).agg(
        次数=('采购单位', 'count'), 销量=('采购数量', 'sum'),
        金额=('成交金额', 'sum'),
    ).sort_values('金额', ascending=False)
    org_stat['金额占比'] = (org_stat['金额'] / total_amt * 100).round(1)

    prices = df['采购单价'].dropna()
    price_stat = pd.DataFrame({
        '指标': ['均价', '最高价', '最低价', '中位数', '价格记录数'],
        '数值': [round(prices.mean(), 1), prices.max(), prices.min(),
                round(prices.median(), 1), len(prices)],
    })

    freq = df.groupby(org_col).size().reset_index(name='采购次数')
    freq.columns = ['采购单位', '采购次数']
    freq_bin = pd.cut(freq['采购次数'], bins=[0, 1, 2, 4, 100],
                      labels=['1次(单次)', '2次', '3-4次', '5次+'])
    freq_stat = freq.groupby(freq_bin, observed=True).size().reset_index(name='单位数')
    freq_stat.columns = ['采购频次区间', '单位数']

    # ============ 各维度订单明细 ============
    detail_brand = make_detail(df[brand_col].notna() if brand_col else pd.Series(True, index=df.index), limit=15) if brand_col else None
    detail_model = make_detail(df.index.isin(df.index), limit=20) if model_col else None
    # 配置价格波动明细：每个波动配置贴「最低价订单 + 最高价订单」对比（谁买贵/买便宜一眼可见）
    detail_stab = None
    if price_stab is not None:
        unstable = price_stab[price_stab['价格状态'] == '⚠️ 有波动']
        if len(unstable) > 0:
            parts = []
            for _, r in unstable.head(5).iterrows():
                sub = df[df['配置'] == r['配置']]
                if len(sub) == 0:
                    continue
                low_p, high_p = sub['采购单价'].min(), sub['采购单价'].max()
                for tag, p in [('最低价', low_p), ('最高价', high_p)]:
                    for _, o in sub[sub['采购单价'] == p].head(1).iterrows():
                        parts.append(pd.DataFrame([{
                            '价格档': tag, '采购单位': o[org_col], '供应商': o[shop_col],
                            '型号': o['型号名'] if model_col else '', '配置': o['配置'] if model_col else '',
                            '采购数量': o['采购数量'], '采购单价(元)': o['采购单价'],
                            '成交时间': o['时间串'],
                        }]))
            if parts:
                detail_stab = pd.concat(parts, ignore_index=True)
    # 多次往来明细
    detail_repeat = None
    if len(repeat) > 0:
        pair_keys = set(zip(repeat['采购单位'], repeat['供应商']))
        detail_repeat = make_detail(df.apply(lambda r: (str(r[org_col]), str(r[shop_col])) in pair_keys, axis=1))
    # TOP采购单位明细：TOP5 单位的全部订单
    detail_org = None
    if len(org_stat) > 0:
        top_orgs = set(org_stat.head(5).index)
        detail_org = make_detail(df[org_col].isin(top_orgs))
    # 峰值月明细
    detail_month = None
    if len(monthly) > 0:
        peak_m = monthly['销量'].idxmax()
        detail_month = make_detail(df['月份'] == peak_m, limit=20)
    # 价格高低订单：最低价和最高价各取若干
    detail_price = None
    if len(prices) > 0:
        low_mask = df['采购单价'] == prices.min()
        high_mask = df['采购单价'] == prices.max()
        detail_price = pd.concat([
            make_detail(low_mask, limit=3), make_detail(high_mask, limit=3)
        ], ignore_index=True)
    # 高频单位明细（3次+）
    detail_freq = None
    high_freq = set(freq[freq['采购次数'] >= 3]['采购单位'])
    if high_freq:
        detail_freq = make_detail(df[org_col].isin(high_freq))
    # 供应商 TOP3 明细
    detail_shop = None
    if len(shop_stat) > 0:
        top_shops = set(shop_stat.head(3).index)
        detail_shop = make_detail(df[shop_col].isin(top_shops), limit=15)

    # ============ 生成说明文本 ============
    def insights_brand():
        if brand_stat is None:
            return ['数据中无品牌字段。']
        lines = [f'共 {len(brand_stat)} 个品牌参与成交。']
        top = brand_stat.iloc[0]
        lines.append(f'主力品牌【{top.name}】：{int(top["销量"])} 台、¥{top["金额"]:,.0f}，金额占比 {top["金额占比"]:.1f}%。')
        if len(brand_stat) > 1:
            cr3 = brand_stat['金额占比'].head(3).sum()
            lines.append(f'TOP3 品牌金额集中度 {cr3:.1f}%' + ('，品牌格局集中。' if cr3 > 70 else '，品牌格局较分散。'))
        lines.append('下方明细为最新成交订单样例。')
        return lines

    def insights_model():
        if model_stat is None:
            return ['数据中无型号字段。']
        lines = [f'共 {len(model_stat)} 个「型号×配置」参与成交。']
        t1 = model_stat.iloc[0]
        cfg = f'（{t1["配置"]}）' if t1.get('配置') else ''
        lines.append(f'最走量款【{str(t1["型号"])[:22]}{cfg}】：{int(t1["销量"])} 台、¥{t1["金额"]:,.0f}、均价 ¥{t1["均价"]:,.0f}。')
        top3 = model_stat.head(3)
        for _, m in top3.iterrows():
            cfg = f'（{m["配置"]}）' if m.get('配置') else ''
            lines.append(f'  · {str(m["型号"])[:18]}{cfg}：{int(m["销量"])} 台、均价 ¥{m["均价"]:,.0f}')
        return lines

    def insights_type():
        lines = [f'共 {len(type_stat)} 类买方单位。']
        t1 = type_stat.iloc[0]
        lines.append(f'成交主力是【{t1.name}】：{int(t1["销量"])} 台、¥{t1["金额"]:,.0f}（{t1["金额占比"]:.1f}%）。')
        for t in ['学校/教育', '医院/医疗']:
            if t in type_stat.index:
                v = type_stat.loc[t, '金额占比']
                if v < 5:
                    lines.append(f'【{t}】金额占比仅 {v:.1f}%，渗透低，是新客户突破口。')
        return lines

    def insights_repeat():
        if len(repeat) == 0:
            return ['未发现同一单位×同一供应商多次成交，均为单次交易（关系尚未固化，正是切入时机）。']
        lines = [f'发现 {len(repeat)} 对固定供货关系，下方明细为每次成交记录：']
        for _, r in repeat.head(5).iterrows():
            lines.append(f'  · {r["采购单位"][:22]} ↔ {r["供应商"][:18]}：{int(r["次数"])} 次、{int(r["销量"])} 台')
        lines.append('这些是已固化的渠道关系；其余单位无固定供货商，是切入机会。')
        return lines

    def insights_shop():
        lines = [f'共 {len(shop_stat)} 家供应商参与供货，下方明细为 TOP3 供应商的最新订单。']
        s1 = shop_stat.iloc[0]
        lines.append(f'头部供应商【{s1.name}】：金额占比 {s1["金额占比"]:.1f}%、客户 {int(s1["客户数"])} 家。')
        cr3 = shop_stat['金额占比'].head(3).sum()
        lines.append(f'TOP3 供应商金额集中度 {cr3:.1f}%' +
                     ('，较集中（头部有壁垒，需从服务/价格切入）。' if cr3 > 60 else '，较分散（无垄断者，切入空间大）。'))
        return lines

    def insights_monthly():
        lines = [f'覆盖 {len(monthly)} 个月的成交，下方明细为销量峰值月的订单。']
        if len(monthly) > 0:
            peak = monthly['销量'].idxmax()
            lines.append(f'销量峰值在【{peak}】：{int(monthly.loc[peak, "销量"])} 台、¥{monthly.loc[peak, "金额"]:,.0f}。')
            growth = monthly['销量'].pct_change()
            if growth.notna().any():
                gmax = growth.idxmax()
                lines.append(f'环比增长最大在【{gmax}】：{growth.loc[gmax]*100:.0f}%。')
        lines.append('若存在明显波峰，采购有周期性，建议在波峰前 1-2 个月做备货与报价准备。')
        return lines

    def insights_org():
        lines = []
        o1 = org_stat.iloc[0]
        lines.append(f'TOP 采购单位【{o1.name[:24]}】：{int(o1["次数"])} 次、{int(o1["销量"])} 台、¥{o1["金额"]:,.0f}（{o1["金额占比"]:.1f}%）。')
        lines.append('下方明细为该单位全部采购记录（供应商/型号/时间），可直接用于客情分析。')
        return lines

    def insights_price():
        lines = [f'成交均价 ¥{prices.mean():,.0f}，中位数 ¥{prices.median():,.0f}。',
                 f'价格区间 ¥{prices.min():,.0f} - ¥{prices.max():,.0f}。',
                 '下方明细为最低价与最高价订单（谁买便宜 / 谁买贵）。']
        return lines

    def insights_freq():
        lines = []
        single = int(freq_stat.loc[freq_stat['采购频次区间'] == '1次(单次)', '单位数'].sum())
        lines.append(f'单次采购单位 {single} 家（占 {single/n_org*100:.0f}%），属"碰运气型"成交。')
        high_n = int(freq_stat.loc[freq_stat['采购频次区间'].isin(['3-4次', '5次+']), '单位数'].sum())
        lines.append(f'高频复购单位（3 次以上）{high_n} 家，下方明细为其采购记录（时间/型号/供应商），值得长期维护。')
        return lines

    def insights_pricestab():
        if price_stab is None:
            return ['数据中无型号字段，跳过。']
        stable_n = int(price_stab['价格状态'].isin(['稳定', '基本稳定(微幅)']).sum())
        lines = [f'共 {len(price_stab)} 个「型号×配置」组合，{stable_n} 个价格稳定。']
        lines.append('价格差异主要来自配置不同（CPU/内存/硬盘/显卡/显示器），属正常。')
        unstable = price_stab[price_stab['价格状态'] == '⚠️ 有波动']
        if len(unstable) > 0:
            lines.append(f'⚠️ {len(unstable)} 个「同一配置价格明显波动」，下方明细为这些配置的全部订单——'
                         f'对比同配置订单，谁买贵了、谁买便宜了，一眼可见：')
            for _, r in unstable.head(3).iterrows():
                cfg = f'（{r["配置"]}）' if r.get('配置') else ''
                lines.append(f'  · {str(r["型号"])[:16]}{cfg}：¥{r["最低价"]:,.0f} - ¥{r["最高价"]:,.0f}'
                             f'（波动 {r["波动幅度"]:,.0f} 元 / {r["波动%"]:.0f}%）')
            lines.append('同配置价格明显波动 = 议价空间或供货差异，是价格谈判的突破口。')
        else:
            lines.append('未发现同配置价格明显波动，市场定价规范。')
        return lines

    # ============ 输出 Excel ============
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # Sheet1 销售明细
        out_parts = []
        if brand_col:
            out_parts.append(('品牌', df[brand_col]))
        if model_col:
            out_parts.append(('型号', df['型号名']))
            out_parts.append(('配置', df['配置']))
        out_parts.extend([
            ('采购单位', df[org_col]), ('供应商', df[shop_col]),
            ('采购数量', df[num_col]), ('采购单价(元)', df[price_col]),
            ('成交时间', df[time_col]),
        ])
        df_out = pd.DataFrame({k: v for k, v in out_parts})
        df_out.to_excel(writer, sheet_name='销售明细', index=False)

        # Sheet2 整体分析报告
        report = gen_overall_report(df, n, total_qty, total_amt, n_org, n_shop,
                                    type_stat, shop_stat, monthly, org_stat, prices,
                                    brand_stat, model_stat, freq_stat, model_col, price_stab)
        ws2 = writer.book.create_sheet('整体分析报告')
        ws2.column_dimensions['A'].width = 110
        for i, line in enumerate(report, 1):
            ws2.cell(row=i, column=1, value=line).font = Font(size=11)
        ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

        # 各分析 Sheet：汇总 + 明细 + 说明
        if brand_stat is not None:
            write_analysis_sheet(writer, '品牌对比分析', brand_stat.reset_index(), detail_brand, insights_brand())
        if model_stat is not None:
            write_analysis_sheet(writer, '型号分析', model_stat, detail_model, insights_model())
        if price_stab is not None:
            write_analysis_sheet(writer, '配置价格稳定性', price_stab, detail_stab, insights_pricestab())
        write_analysis_sheet(writer, '单位类型分布', type_stat.reset_index(), None, insights_type())
        write_analysis_sheet(writer, '多次往来', repeat.reset_index(drop=True), detail_repeat, insights_repeat())
        write_analysis_sheet(writer, '供应商占比', shop_stat.reset_index(), detail_shop, insights_shop())
        write_analysis_sheet(writer, '月度销售趋势', monthly.reset_index(), detail_month, insights_monthly())
        write_analysis_sheet(writer, 'TOP采购单位', org_stat.reset_index(), detail_org, insights_org())
        write_analysis_sheet(writer, '价格分析', price_stat, detail_price, insights_price())
        write_analysis_sheet(writer, '采购频次', freq_stat, detail_freq, insights_freq())

    # 控制台摘要（带关键订单）—— --quiet 时跳过，只留完成行
    if not quiet:
        print('\n■ 主要洞察：')
        for fn in (insights_brand, insights_type, insights_shop, insights_repeat):
            for line in fn():
                print('  ' + line)
            print()

        # 控制台：价格波动订单（最低价 vs 最高价对比）
        if detail_stab is not None and len(detail_stab) > 0:
            print('■ ⚠️ 同配置价格波动（最低价 vs 最高价，议价空间）：')
            for _, r in detail_stab.iterrows():
                tag = '🔻最低' if r['价格档'] == '最低价' else '🔺最高'
                print(f'  {tag} {str(r["型号"])[:14]}（{str(r["配置"])[:22]}）| {r["采购单位"][:20]} | '
                      f'{r["供应商"][:14]} | {int(r["采购数量"])}台 ¥{r["采购单价(元)"]:,.0f} | {r["成交时间"]}')
            print()

    print(f'\n分析完成，报告已保存：{out_path}')
    with pd.ExcelFile(out_path) as xf:
        print(f'  Sheet：{xf.sheet_names}')


def main():
    parser = argparse.ArgumentParser(description='京华云采销售数据分析（数据+订单明细+说明）')
    parser.add_argument('input', help='销售记录 xlsx（fetch_sale_records.py 的输出）')
    parser.add_argument('--out', default=None, help='分析报告输出路径（默认 原名_分析报告.xlsx）')
    parser.add_argument('--quiet', action='store_true', help='安静模式：不打印洞察明细，只打印完成行（省 token）')
    args = parser.parse_args()

    input_path = os.path.abspath(args.input)
    if not os.path.exists(input_path):
        print(f'❌ 文件不存在：{input_path}')
        sys.exit(1)

    out_path = args.out or input_path.rsplit('.', 1)[0] + '_分析报告.xlsx'
    analyze(input_path, out_path, quiet=args.quiet)


if __name__ == '__main__':
    main()
