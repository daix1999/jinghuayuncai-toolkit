#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
京华云采经销商爬取脚本（按产品 / 品类 / 品牌维度）
正向爬取经销商及其联系方式，统一输出格式：品类、品牌、产品、经销商名称、电话。

用法：
  # 指定品类下所有产品的经销商
  python crawl_dealers.py --category 台 --out 台式机经销商.xlsx

  # 指定品牌下所有产品的经销商
  python crawl_dealers.py --brand 联想 --out 联想经销商.xlsx

  # 指定产品（关键词）的经销商
  python crawl_dealers.py --product 开天M70d --out 开天M70d经销商.xlsx

  # 组合：台式品类 + 联想品牌
  python crawl_dealers.py --category 台 --brand 联想 --out 台式联想经销商.xlsx

  # 多个品类、多个品牌可逗号分隔
  python crawl_dealers.py --category 台,笔,服 --brand 联想,华为 --out 多品类多品牌.xlsx

筛选说明：
  - --category：品类键（台/笔/服/印），不指定则全部品类
  - --brand：品牌名，按 brandName 包含匹配（模糊，如"联想"可匹配"联想"）
  - --product：产品关键词，按 skuName 包含匹配（如"开天"、"M70d"、"X5z"）
"""
import argparse
import json
import os
import time
import requests
from collections import defaultdict

BASE = "https://mkt-bjzc.zhongcy.com"
SEARCH_URL = BASE + "/proxy/trade-service/mall/search/querySkuListFromEs"
AGENT_URL = BASE + "/proxy/trade-service/mall/search/querySkuAgentListFromEs"
PLATFORM_ID = 20
PUBLISH_TYPE = 10024

# 品类键 -> (cid, 名称)
CATEGORIES = {
    "台": (1000011, "台式计算机"),
    "笔": (1000014, "便携式计算机"),
    "服": (1000128, "服务器"),
    "印": (1000023, "打印机"),  # 打印机有多个子 cid
}

# 打印机扩展 cid（喷墨/激光/针式/多功能/复印等）
PRINTER_CIDS = [1000023, 1000025, 1000027, 1000029, 1000041]

SESSION = requests.Session()
# 绕过系统代理直连（避免代理导致 zhongcy.com TLS 握手失败）
SESSION.trust_env = False
SESSION.headers.update({
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": BASE + "/mall-view/product/search",
    "Origin": BASE,
})


def post_json(url, payload, retries=3):
    for i in range(retries):
        try:
            r = SESSION.post(url, json=payload, timeout=30)
            d = r.json()
            if d.get("code") == "0":
                return d
            if "99999" in str(d.get("code", "")):
                time.sleep(0.5 * (i + 2))
                continue
            return d
        except Exception:
            time.sleep(0.5 * (i + 2))
    return None


def get_cid_list(category_keys):
    """品类键 -> cid 列表（打印机展开多个子 cid）"""
    if not category_keys:
        # 全部品类
        cids = []
        for k, (cid, _) in CATEGORIES.items():
            if k == "印":
                cids.extend(PRINTER_CIDS)
            else:
                cids.append(cid)
    else:
        cids = []
        for k in category_keys:
            if k == "印":
                cids.extend(PRINTER_CIDS)
            elif k in CATEGORIES:
                cids.append(CATEGORIES[k][0])
    # 去重保序
    seen, result = set(), []
    for c in cids:
        if c not in seen:
            seen.add(c)
            result.append(c)
    return result


def cid_to_name(cid):
    for k, (c, name) in CATEGORIES.items():
        if c == cid:
            return name
    return str(cid)


def get_products(cid):
    """获取某 cid 下所有商品"""
    products = []
    page, page_size = 1, 100
    while True:
        payload = {
            "queryPage": {"platformId": PLATFORM_ID, "pageSize": page_size, "pageNum": page},
            "homeType": 110,
            "isAggregation": False,
            "isHometypes": True,
            "cid": cid,
            "businessType": 1,
            "publishType": PUBLISH_TYPE,
        }
        d = post_json(SEARCH_URL, payload)
        if not d or not d.get("data"):
            break
        rl = d["data"]["itemList"]["resultList"]
        if not rl:
            break
        for it in rl:
            products.append({
                "skuId": it["skuId"],
                "skuName": it.get("skuName", ""),
                "brandName": (it.get("brandNameCh") or "") + (it.get("brandNameEn") or ""),
                "category": cid_to_name(cid),
            })
        total = int(d["data"].get("totalCount", 0) or 0)
        if page * page_size >= total:
            break
        page += 1
        time.sleep(0.2)
    return products


def get_dealers(sku_id):
    """获取某商品全部经销商（含 agentPhone）"""
    dealers = []
    page, page_size = 1, 100  # 上限 100，超限服务端排序异常
    while True:
        payload = {
            "platformId": PLATFORM_ID,
            "skuId": sku_id,
            "shopName": "",
            "queryPage": {"platformId": PLATFORM_ID, "pageSize": page_size, "pageNum": page},
            "publishType": PUBLISH_TYPE,
        }
        d = post_json(AGENT_URL, payload)
        if not d or not d.get("data"):
            break
        ia = d["data"]["itemAgentList"]
        rl = ia.get("resultList", [])
        dealers.extend(rl)
        total = ia.get("count", 0)
        if page * page_size >= total or not rl:
            break
        page += 1
        time.sleep(0.25)
    return dealers


def filter_products(products, brands, product_kw):
    """按品牌、产品关键词筛选商品"""
    result = []
    for p in products:
        if brands and not any(b in p["brandName"] for b in brands):
            continue
        if product_kw and product_kw not in p["skuName"]:
            continue
        result.append(p)
    return result


def export_excel(rows, out_path):
    """输出统一格式 Excel：品类、品牌、产品、经销商名称、电话"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "经销商列表"

    headers = ["品类", "品牌", "产品", "经销商名称", "销售电话"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2F5496")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)

    widths = [14, 12, 30, 36, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"输出：{out_path}（共 {len(rows)} 条经销商记录）")


def main():
    parser = argparse.ArgumentParser(description="京华云采经销商爬取（按产品/品类/品牌）")
    parser.add_argument("--category", default="", help="品类键（台/笔/服/印），逗号分隔，默认全部")
    parser.add_argument("--brand", default="", help="品牌名（如 联想），逗号分隔，模糊匹配")
    parser.add_argument("--product", default="", help="产品关键词（如 开天M70d），按 skuName 包含匹配")
    parser.add_argument("--out", default="经销商列表.xlsx", help="输出 Excel 路径")
    args = parser.parse_args()

    cat_keys = [k.strip() for k in args.category.split(",") if k.strip()]
    brands = [b.strip() for b in args.brand.split(",") if b.strip()]
    product_kw = args.product.strip()

    print(f"筛选条件：品类={cat_keys or '全部'}，品牌={brands or '全部'}，产品={product_kw or '全部'}")

    cids = get_cid_list(cat_keys)
    print(f"遍历 cid：{cids}")

    all_products = []
    for cid in cids:
        prods = get_products(cid)
        all_products.extend(prods)
        print(f"cid={cid}（{cid_to_name(cid)}）：{len(prods)} 个商品")

    # 筛选
    filtered = filter_products(all_products, brands, product_kw)
    print(f"\n筛选后共 {len(filtered)} 个商品，开始爬取经销商...")

    # 聚合：经销商名 -> 集合
    dealer_map = defaultdict(set)  # dealer_name -> set of (category, brand, product, phone)

    for idx, p in enumerate(filtered, 1):
        dealers = get_dealers(p["skuId"])
        for dl in dealers:
            aname = (dl.get("agentName") or "").strip()
            if not aname:
                continue
            phone = dl.get("agentPhone", "") or ""
            dealer_map[aname].add((p["category"], p["brandName"], p["skuName"], str(phone)))
        if idx % 20 == 0:
            print(f"  [{idx}/{len(filtered)}] 已发现 {len(dealer_map)} 个经销商")
        time.sleep(0.15)

    # 展开成行
    rows = []
    for aname, records in sorted(dealer_map.items()):
        # 按品类、品牌、产品排序，聚合
        cats = sorted({r[0] for r in records})
        brands = sorted({r[1] for r in records})
        products = sorted({r[2] for r in records})
        phones = sorted({r[3] for r in records if r[3]})
        rows.append([
            "、".join(cats),
            "、".join(brands),
            "、".join(products),
            aname,
            " / ".join(phones),
        ])

    print(f"\n爬取完成：{len(rows)} 个经销商")
    export_excel(rows, args.out)


if __name__ == "__main__":
    main()
